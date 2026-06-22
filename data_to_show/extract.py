import openpyxl
import glob
import re
import json
import os
from collections import defaultdict

def parse_benefits(text):
    result = {
        'dinar_qty': 0,
        'gold_qty': 0,
        'gold_grams': 0,
        'membership_qty': 0,
        'card_qty': 0,
        'tags': []
    }
    if not text or text == 'N/A':
        return result

    t = text.upper().strip()
    tags = set()

    # VAQUITA: shared purchase pool
    if 'VAQUITA' in t:
        vaquita = re.findall(r'EQUIVALENTE\s*A\s*(\d+)\s*CAJAS?\s*(?:DE\s*)?DINAR', t)
        if vaquita:
            result['dinar_qty'] = sum(int(x) for x in vaquita)
            tags.add('dinar')
        tags.add('vaquita')

    # COMPRA COMPARTIDA: shared purchase with receive amount
    elif 'COMPRA COMPARTIDA' in t or 'RECIBO' in t:
        compartida = re.search(r'RECIBO\s*(\d+)\s*DE', t)
        if compartida:
            result['dinar_qty'] = int(compartida.group(1))
            tags.add('dinar')

    # Standard: catch all "X CAJAS DINARES IRAKIES ROJOS" patterns
    else:
        dinar_boxes = re.findall(
            r'(\d+)\s*CAJAS?\s*(?:DE\s*)?DINAR(?:ES)?\s*(?:IRAKIES|ROJOS|IRAQUIES)', t
        )
        if dinar_boxes:
            result['dinar_qty'] = sum(int(x) for x in dinar_boxes)
            tags.add('dinar')

    # CAJAS DE DINARES without a leading number (e.g. "CAJAS DE DINARES X 40.000 NOTAS")
    if re.search(r'CAJAS?\s*(?:DE\s*)?DINAR', t) and result['dinar_qty'] == 0:
        tags.add('dinar')

    # GOLD MICROLINGOTS
    gold_boxes = re.findall(r'(\d+)\s*CAJAS?\s*(?:DE\s*)?MICROLINGOTE', t)
    if gold_boxes:
        result['gold_qty'] = sum(int(x) for x in gold_boxes)
        tags.add('gold')

    # Gold without number prefix (e.g. "CAJA DE MICROLINGOTES POR 500")
    if re.search(r'CAJA\s*(?:DE\s*)?MICROLINGOTE', t) and not gold_boxes:
        result['gold_qty'] += 1
        tags.add('gold')

    # CONTENEDOR DE ORO
    if re.search(r'CONTENEDOR\s*DE\s*ORO', t):
        cajas = re.search(r'([\d,]+)\s*CAJAS', t)
        if cajas:
            result['gold_qty'] += int(cajas.group(1).replace(',', ''))
        tags.add('gold')

    # GOLD GRAMS: parse GR weight from MICROLINGOTE/ORO context
    if 'gold' in tags or 'MICROLINGOTE' in t or 'ORO' in t:
        gr_match = re.search(r'(\d+)\s*GR', t)
        if gr_match:
            result['gold_grams'] = int(gr_match.group(1))

    # MEMBERSHIPS
    membresias = re.findall(r'(\d+)\s*MEMBRESIA', t)
    if not membresias:
        membresias = re.findall(r'(\d+)\s*MEMBRECIA', t)  # typo variant
    if membresias:
        result['membership_qty'] = sum(int(x) for x in membresias)
        tags.add('membership')

    # CARDS
    tarjetas = re.findall(r'(\d+)\s*TARJETA', t)
    if tarjetas:
        result['card_qty'] = sum(int(x) for x in tarjetas)
        tags.add('card')

    # BONUS
    if re.search(r'BONO|BONUS', t):
        tags.add('bonus')

    if not tags:
        tags.add('other')

    result['tags'] = sorted(tags)
    return result


def extract_personal_data(ws):
    data = {}
    # Read label-value pairs: label is in column A, value in column B
    label_map = {
        3: 'name',
        4: 'cedula',
        5: 'email',
        6: 'telegram',
        7: 'phone',
        8: 'birthdate',
    }
    for row_num, key in label_map.items():
        val = ws.cell(row_num, 2).value
        if val:
            data[key] = str(val).strip()
        else:
            data[key] = ''

    # Location info
    for row_num, key in [(16, 'country'), (17, 'department'), (18, 'city'), (19, 'zip'), (20, 'address')]:
        val = ws.cell(row_num, 2).value
        if val:
            data[key] = str(val).strip()
        else:
            data[key] = ''

    return data


def normalize_flayer(name):
    if not name:
        return ''
    n = name.strip()
    n = re.sub(r'\s+', ' ', n)
    return n


def main():
    files = sorted(glob.glob('db/*.xlsx'))
    members = []
    all_purchases = []

    for fpath in files:
        wb = openpyxl.load_workbook(fpath, data_only=True)
        fname = os.path.basename(fpath)

        # Extract personal data
        ws1 = wb['DATOS PERSONALES']
        personal = extract_personal_data(ws1)

        # Use filename as fallback for name
        if not personal.get('name'):
            name_from_file = fname \
                .replace('DATOS_PERSONALES_Y_TABLA_DE_HISTORIAL_DE_COMPRAS_', '') \
                .replace('.xlsx', '') \
                .replace('_', ' ') \
                .strip()
            personal['name'] = name_from_file

        member_id = personal.get('cedula', '') or personal.get('name', '')

        # Extract purchase records
        ws2 = wb['HISTORIAL APORTES']
        purchases = []
        for r in range(4, ws2.max_row + 1):
            seq = ws2.cell(r, 1).value
            if seq is None:
                continue

            flayer = normalize_flayer(str(ws2.cell(r, 9).value or ''))
            benef_text = str(ws2.cell(r, 14).value or '').strip()
            amount_raw = ws2.cell(r, 8).value
            amount = None
            if amount_raw and amount_raw != 'N/A':
                try:
                    amount = float(amount_raw)
                except (ValueError, TypeError):
                    pass

            date_val = ws2.cell(r, 4).value
            date_str = ''
            if date_val:
                try:
                    date_str = date_val.strftime('%Y-%m-%d')
                except AttributeError:
                    date_str = str(date_val)

            year_val = ws2.cell(r, 5).value
            year = int(year_val) if year_val else None

            payment = str(ws2.cell(r, 10).value or '').strip()
            leader = str(ws2.cell(r, 15).value or '').strip()
            amount_to_receive = str(ws2.cell(r, 13).value or '').strip()

            # We need to handle member name from the purchase row too
            purchase_name = str(ws2.cell(r, 2).value or '').strip()

            parsed = parse_benefits(benef_text)

            # VAQUITA: Column H has the full pot ($6M for everyone).
            # Real individual share is in benefit text: "CONSIGNACION $50,000"
            if 'vaquita' in parsed['tags']:
                consignacion = re.search(
                    r'CONSIGNACION\s*\$?\s*([0-9,]+)', benef_text.upper()
                )
                if consignacion:
                    try:
                        amount = float(consignacion.group(1).replace(',', ''))
                    except (ValueError, TypeError):
                        pass

            purchase = {
                'date': date_str,
                'year': year,
                'flayer': flayer,
                'amount': amount,
                'payment': payment,
                'benefit_text': benef_text,
                'dinar_qty': parsed['dinar_qty'],
                'gold_qty': parsed['gold_qty'],
                'gold_grams': parsed['gold_grams'],
                'membership_qty': parsed['membership_qty'],
                'card_qty': parsed['card_qty'],
                'tags': parsed['tags'],
                'leader': leader,
                'amount_to_receive': amount_to_receive,
            }
            purchases.append(purchase)
            all_purchases.append(purchase)

        member = {
            'id': member_id,
            'name': personal.get('name', ''),
            'cedula': personal.get('cedula', ''),
            'email': personal.get('email', ''),
            'telegram': personal.get('telegram', ''),
            'phone': personal.get('phone', ''),
            'birthdate': personal.get('birthdate', ''),
            'country': personal.get('country', ''),
            'department': personal.get('department', ''),
            'city': personal.get('city', ''),
            'purchases': purchases,
            'purchase_count': len(purchases),
            'total_spent': sum(p['amount'] for p in purchases if p['amount']),
        }
        members.append(member)

    # Build summary stats
    total_collected = sum(m['total_spent'] for m in members)
    total_purchases = sum(m['purchase_count'] for m in members)

    # Category totals
    cat_totals = defaultdict(int)
    flayer_counts = defaultdict(int)
    flayer_details = defaultdict(lambda: {
        'dinar_qty': 0, 'gold_qty': 0, 'gold_grams': set(), 'membership_qty': 0, 'card_qty': 0,
        'tags': set(), 'purchase_count': 0, 'total_amount': 0,
        'benefit_samples': []
    })

    for p in all_purchases:
        for tag in p['tags']:
            cat_totals[tag] += 1
        flayer_counts[p['flayer']] += 1
        fd = flayer_details[p['flayer']]
        fd['dinar_qty'] += p['dinar_qty']
        fd['gold_qty'] += p['gold_qty']
        if p['gold_grams']:
            fd['gold_grams'].add(p['gold_grams'])
        fd['membership_qty'] += p['membership_qty']
        fd['card_qty'] += p['card_qty']
        fd['tags'].update(p['tags'])
        fd['purchase_count'] += 1
        if p['amount']:
            fd['total_amount'] += p['amount']
        if p['benefit_text'] and len(fd['benefit_samples']) < 3:
            fd['benefit_samples'].append(p['benefit_text'])

    # Convert flayer_details to serializable format
    flayers_list = []
    for name, fd in sorted(flayer_details.items()):
        flayers_list.append({
            'name': name,
            'count': flayer_counts[name],
            'dinar_qty': fd['dinar_qty'],
            'gold_qty': fd['gold_qty'],
            'gold_grams': sorted(fd['gold_grams']),
            'membership_qty': fd['membership_qty'],
            'card_qty': fd['card_qty'],
            'tags': sorted(fd['tags']),
            'total_amount': fd['total_amount'],
            'benefit_samples': fd['benefit_samples'],
        })

    summary = {
        'total_members': len(members),
        'total_purchases': total_purchases,
        'total_collected': total_collected,
        'category_totals': dict(cat_totals),
        'flayers': flayers_list,
    }

    output = {
        'summary': summary,
        'members': members,
    }

    with open('dashboard-data.json', 'w', encoding='utf-8') as f:
        json.dump(output, f, ensure_ascii=False, indent=2)

    print(f'Done! Extracted {total_purchases} purchases from {len(members)} members')
    print(f'Total collected: ${total_collected:,.0f}')
    print(f'Unique flayer types: {len(flayers_list)}')
    print(f'Category totals: {dict(cat_totals)}')
    print(f'Output: dashboard-data.json')


if __name__ == '__main__':
    main()

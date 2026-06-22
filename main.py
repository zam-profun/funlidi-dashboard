import os
import io
import json
import re
import secrets
import time
import difflib
from datetime import datetime, timezone, timedelta
from collections import defaultdict

import openpyxl

from dotenv import load_dotenv
from fastapi import FastAPI, Body, HTTPException, Request
from fastapi.responses import FileResponse, StreamingResponse, JSONResponse
from fastapi.staticfiles import StaticFiles
from supabase import create_client, Client
from openpyxl import Workbook

load_dotenv()

SUPABASE_URL = os.environ["SUPABASE_URL"]
SUPABASE_KEY = os.environ["SUPABASE_SERVICE_KEY"]
AYUDAS_SUPABASE_URL = os.environ.get("AYUDAS_SUPABASE_URL") or os.environ["SUPABASE_URL"]
AYUDAS_SUPABASE_KEY = os.environ.get("AYUDAS_SUPABASE_SERVICE_KEY") or os.environ["SUPABASE_SERVICE_KEY"]
INVENTARIO_SUPABASE_URL = os.environ.get("INVENTARIO_SUPABASE_URL") or os.environ["SUPABASE_URL"]
INVENTARIO_SUPABASE_KEY = os.environ.get("INVENTARIO_SUPABASE_SERVICE_KEY") or os.environ["SUPABASE_SERVICE_KEY"]

# --- Auth ---
_USERS = {}
if os.getenv("PASSWORD_MARIA"):
    _USERS["Maria"] = os.environ["PASSWORD_MARIA"]
if os.getenv("PASSWORD_JEOVANI"):
    _USERS["Jeovani"] = os.environ["PASSWORD_JEOVANI"]
if os.getenv("PASSWORD_ANGEL"):
    _USERS["Angel"] = os.environ["PASSWORD_ANGEL"]

_SESSIONS = {}  # token -> {"expiry": float, "username": str}

def _clean_sessions():
    now = time.time()
    expired = [k for k, v in _SESSIONS.items() if v["expiry"] < now]
    for k in expired:
        del _SESSIONS[k]
# --- end Auth ---

COL_TZ = timezone(timedelta(hours=-5))

DATA_DIR = os.path.join(os.path.dirname(os.path.abspath(__file__)), "data_to_show")
_PAGOS_CACHE = {"data": None, "mtime": 0}
_CRM_CACHE = {"data": None, "mtime": 0}

PAGOS_FILES = [
    "EXCEL ACTIVACION PAGOS MARIA ELVIRA SUS.xlsx",
    "EXCEL INFINITY - LOGISTICA - ARITA MARIA ELVIRA SUS.xlsx",
]


def _load_pagos_data():
    latest_mtime = 0
    for fname in PAGOS_FILES:
        fpath = os.path.join(DATA_DIR, fname)
        if os.path.exists(fpath):
            mtime = os.path.getmtime(fpath)
            if mtime > latest_mtime:
                latest_mtime = mtime
    if latest_mtime == 0:
        return []

    if _PAGOS_CACHE["data"] is not None and latest_mtime <= _PAGOS_CACHE["mtime"]:
        return _PAGOS_CACHE["data"]

    records = []
    for fname in PAGOS_FILES:
        fpath = os.path.join(DATA_DIR, fname)
        if not os.path.exists(fpath):
            continue
        wb = openpyxl.load_workbook(fpath, read_only=True, data_only=True)
        ws = wb[wb.sheetnames[0]]
        for row in ws.iter_rows(min_row=3, values_only=True):
            if row[0] is None and row[1] is None:
                break
            num, ident, nombres, ref, fecha, hora, flayer, valor = row[0:8]
            fecha_str = ""
            if fecha:
                try:
                    fecha_str = fecha.isoformat() if hasattr(fecha, "isoformat") else str(fecha)
                except Exception:
                    fecha_str = str(fecha)
            hora_str = ""
            if hora:
                try:
                    hora_str = hora.strftime("%H:%M") if hasattr(hora, "strftime") else str(hora)
                except Exception:
                    hora_str = str(hora)
            records.append({
                "id": int(num) if num else 0,
                "identificacion": str(ident).strip() if ident else "",
                "nombres": str(nombres).strip() if nombres else "",
                "referencia": str(ref).strip() if ref else "",
                "fecha": fecha_str,
                "hora": hora_str,
                "flayer": re.sub(r'\s+', ' ', str(flayer)).strip() if flayer else "",
                "valor": int(valor) if isinstance(valor, (int, float)) else 0,
                "archivo": fname,
            })
        wb.close()

    _PAGOS_CACHE["data"] = records
    _PAGOS_CACHE["mtime"] = latest_mtime
    return records


def _load_crm_data():
    fpath = os.path.join(DATA_DIR, "dashboard-data.json")
    if not os.path.exists(fpath):
        return None
    mtime = os.path.getmtime(fpath)
    if _CRM_CACHE["data"] is not None and mtime <= _CRM_CACHE["mtime"]:
        return _CRM_CACHE["data"]
    with open(fpath, "r", encoding="utf-8") as f:
        data = json.load(f)
    _CRM_CACHE["data"] = data
    _CRM_CACHE["mtime"] = mtime
    return data


supabase: Client = create_client(SUPABASE_URL, SUPABASE_KEY)
supabase_ayudas: Client = create_client(AYUDAS_SUPABASE_URL, AYUDAS_SUPABASE_KEY)
supabase_inventario: Client = create_client(INVENTARIO_SUPABASE_URL, INVENTARIO_SUPABASE_KEY)

app = FastAPI()

app.mount("/static", StaticFiles(directory="static"), name="static")


@app.post("/api/auth/login")
async def auth_login(data: dict):
    pwd = data.get("password", "")
    username = None
    for u, pw in _USERS.items():
        if pw == pwd:
            username = u
            break
    if not username:
        raise HTTPException(401, "Contraseña incorrecta")
    token = secrets.token_hex(32)
    _SESSIONS[token] = {"expiry": time.time() + 86400, "username": username}
    resp = JSONResponse({"ok": True, "username": username})
    resp.set_cookie(key="session", value=token, httponly=True,
                    samesite="lax", max_age=86400)
    return resp


@app.post("/api/auth/logout")
async def auth_logout(request: Request):
    token = request.cookies.get("session")
    _SESSIONS.pop(token, None)
    resp = JSONResponse({"ok": True})
    resp.delete_cookie("session")
    return resp


@app.get("/api/auth/check")
async def auth_check(request: Request):
    token = request.cookies.get("session")
    _clean_sessions()
    if token and token in _SESSIONS:
        return {"authenticated": True, "username": _SESSIONS[token]["username"]}
    return {"authenticated": False}


@app.middleware("http")
async def session_middleware(request: Request, call_next):
    path = request.url.path
    if path.startswith("/api/") and not path.startswith("/api/auth/"):
        token = request.cookies.get("session")
        _clean_sessions()
        if not token or token not in _SESSIONS:
            return JSONResponse(status_code=401, content={"detail": "No autorizado"})
    return await call_next(request)


@app.get("/")
async def index():
    return FileResponse("static/index.html")


@app.get("/api/data")
async def get_data():
    result = supabase.table("usuarios_funlidi").select("*").order("updated_at", desc=True).execute()
    return {"data": result.data, "total": len(result.data)}


@app.get("/api/stats")
async def get_stats():
    result = supabase.table("usuarios_funlidi").select("*").execute()
    rows = result.data
    total = len(rows)
    completados = sum(
        1 for r in rows
        if r.get("nombres_completos") and r.get("correo_electronico") and r.get("numero_documento")
    )
    incompletos = total - completados
    ultima = max(
        (r.get("updated_at") or r.get("created_at") or "") for r in rows
    ) if rows else None

    ahora_col = datetime.now(COL_TZ)
    hoy_inicio_col = ahora_col.replace(hour=0, minute=0, second=0, microsecond=0)
    semana_inicio_col = hoy_inicio_col - timedelta(days=7)

    registros_hoy = 0
    actualizaciones_hoy = 0
    registros_semana = 0
    actualizaciones_semana = 0

    for r in rows:
        creado = r.get("created_at")
        actualizado = r.get("updated_at")
        if creado:
            try:
                c = datetime.fromisoformat(creado.replace("Z", "+00:00")).astimezone(COL_TZ)
                if c >= hoy_inicio_col:
                    registros_hoy += 1
                if c >= semana_inicio_col:
                    registros_semana += 1
            except Exception:
                pass
        if actualizado:
            try:
                a = datetime.fromisoformat(actualizado.replace("Z", "+00:00")).astimezone(COL_TZ)
                if a >= hoy_inicio_col:
                    if creado:
                        try:
                            c = datetime.fromisoformat(creado.replace("Z", "+00:00")).astimezone(COL_TZ)
                            if abs((a - c).total_seconds()) > 5:
                                actualizaciones_hoy += 1
                        except Exception:
                            actualizaciones_hoy += 1
                    else:
                        actualizaciones_hoy += 1
                if a >= semana_inicio_col:
                    if creado:
                        try:
                            c = datetime.fromisoformat(creado.replace("Z", "+00:00")).astimezone(COL_TZ)
                            if abs((a - c).total_seconds()) > 5:
                                actualizaciones_semana += 1
                        except Exception:
                            actualizaciones_semana += 1
                    else:
                        actualizaciones_semana += 1
            except Exception:
                pass

    return {
        "total": total,
        "completados": completados,
        "incompletos": incompletos,
        "ultima_actualizacion": ultima,
        "registros_hoy": registros_hoy,
        "actualizaciones_hoy": actualizaciones_hoy,
        "registros_semana": registros_semana,
        "actualizaciones_semana": actualizaciones_semana,
    }


@app.get("/api/activity")
async def get_activity():
    result = supabase.table("usuarios_funlidi").select("*").order("updated_at", desc=True).limit(10).execute()
    rows = result.data
    actividades = []
    for r in rows:
        creado = r.get("created_at")
        actualizado = r.get("updated_at")
        tipo = "Nuevo registro"
        if creado and actualizado:
            try:
                c = datetime.fromisoformat(creado.replace("Z", "+00:00"))
                a = datetime.fromisoformat(actualizado.replace("Z", "+00:00"))
                if abs((a - c).total_seconds()) > 5:
                    tipo = "Informacion actualizada"
            except Exception:
                pass
        timestamp = actualizado or creado
        usuario = r.get("telegram_username")
        if usuario:
            usuario = "@" + usuario
        else:
            usuario = str(r.get("telegram_user_id", ""))
        nombre = r.get("nombres_completos") or "(sin nombre)"
        actividades.append({
            "usuario": usuario,
            "nombre": nombre,
            "tipo": tipo,
            "timestamp": timestamp,
        })
    return {"actividades": actividades}


@app.get("/api/anomalies")
async def get_anomalies():
    result = supabase.table("usuarios_funlidi").select("*").execute()
    rows = result.data
    anomalias = []

    usuarios = {}
    for r in rows:
        uid = r.get("telegram_user_id")
        if uid is not None:
            usuarios[uid] = r

    docs = {}
    correos = {}
    nombres = []

    for uid, r in usuarios.items():
        doc = r.get("numero_documento")
        corr = r.get("correo_electronico")
        nom = r.get("nombres_completos")
        usr = r.get("telegram_username") or str(uid)
        if usr and not usr.startswith("@"):
            usr = "@" + usr

        if doc:
            docs.setdefault(doc, []).append({"uid": uid, "usuario": usr, "nombre": nom})
        if corr:
            correos.setdefault(corr, []).append({"uid": uid, "usuario": usr, "nombre": nom})
        if nom:
            nombres.append({"uid": uid, "usuario": usr, "nombre": nom})

    for doc, involucrados in docs.items():
        if len(involucrados) > 1:
            uids = list(set(i["uid"] for i in involucrados))
            if len(uids) > 1:
                anomalias.append({
                    "tipo": "documento",
                    "descripcion": f"El numero de documento {doc} esta siendo usado por {len(uids)} personas distintas.",
                    "involucrados": [{"usuario": i["usuario"], "nombre": i["nombre"] or "(sin nombre)"} for i in involucrados],
                })

    for corr, involucrados in correos.items():
        if len(involucrados) > 1:
            uids = list(set(i["uid"] for i in involucrados))
            if len(uids) > 1:
                anomalias.append({
                    "tipo": "correo",
                    "descripcion": f"El correo electronico {corr} esta siendo usado por {len(uids)} personas distintas.",
                    "involucrados": [{"usuario": i["usuario"], "nombre": i["nombre"] or "(sin nombre)"} for i in involucrados],
                })

    for i in range(len(nombres)):
        for j in range(i + 1, len(nombres)):
            if nombres[i]["uid"] == nombres[j]["uid"]:
                continue
            a = (nombres[i]["nombre"] or "").upper().strip()
            b = (nombres[j]["nombre"] or "").upper().strip()
            if not a or not b:
                continue
            ratio = difflib.SequenceMatcher(None, a, b).ratio()
            if ratio >= 0.75:
                anomalias.append({
                    "tipo": "nombre",
                    "descripcion": (
                        f"Los nombres son muy similares ({int(ratio * 100)}% de coincidencia): "
                        f'"{nombres[i]["nombre"]}" y "{nombres[j]["nombre"]}"'
                    ),
                    "involucrados": [
                        {"usuario": nombres[i]["usuario"], "nombre": nombres[i]["nombre"] or "(sin nombre)"},
                        {"usuario": nombres[j]["usuario"], "nombre": nombres[j]["nombre"] or "(sin nombre)"},
                    ],
                })

    return {"anomalias": anomalias, "total": len(anomalias)}


@app.get("/api/pagos/data")
async def get_pagos_data():
    records = _load_pagos_data()
    return {"data": records, "total": len(records)}


@app.get("/api/pagos/stats")
async def get_pagos_stats():
    records = _load_pagos_data()
    total_cop = sum(r["valor"] for r in records)
    total_trans = len(records)
    ids_unicos = set(r["identificacion"] for r in records if r["identificacion"])
    total_personas = len(ids_unicos)

    ultima = ""
    for r in records:
        if r["fecha"] and r["fecha"] > ultima:
            ultima = r["fecha"]

    por_flayer = defaultdict(lambda: {"cantidad": 0, "total_cop": 0, "personas": set()})
    for r in records:
        f = r["flayer"] or "SIN ESPECIFICAR"
        por_flayer[f]["cantidad"] += 1
        por_flayer[f]["total_cop"] += r["valor"]
        if r["identificacion"]:
            por_flayer[f]["personas"].add(r["identificacion"])

    flayer_list = []
    for f, d in sorted(por_flayer.items(), key=lambda x: -x[1]["total_cop"]):
        flayer_list.append({
            "flayer": f,
            "cantidad": d["cantidad"],
            "total_cop": d["total_cop"],
            "personas_unicas": len(d["personas"]),
            "porcentaje_cop": round(d["total_cop"] / total_cop * 100, 1) if total_cop else 0,
        })

    por_dia = defaultdict(lambda: {"cantidad": 0, "total_cop": 0})
    for r in records:
        if not r["fecha"]:
            continue
        dia = r["fecha"][:10]
        por_dia[dia]["cantidad"] += 1
        por_dia[dia]["total_cop"] += r["valor"]

    dia_list = [{"fecha": d, **v} for d, v in sorted(por_dia.items())]

    return {
        "total_cop": total_cop,
        "total_transacciones": total_trans,
        "total_personas_unicas": total_personas,
        "ultima_transaccion": ultima,
        "por_flayer": flayer_list,
        "por_dia": dia_list,
    }


@app.get("/api/pagos/personas")
async def get_pagos_personas():
    records = _load_pagos_data()
    personas = defaultdict(lambda: {
        "nombres": "", "total_gastado": 0, "transacciones": 0,
        "flyers": set(), "referencias": [], "primer_pago": "", "ultimo_pago": "",
    })
    for r in records:
        ident = r["identificacion"]
        if not ident:
            continue
        p = personas[ident]
        if not p["nombres"]:
            p["nombres"] = r["nombres"]
        p["total_gastado"] += r["valor"]
        p["transacciones"] += 1
        p["flyers"].add(r["flayer"])
        p["referencias"].append(r["referencia"])
        if r["fecha"]:
            if not p["primer_pago"] or r["fecha"] < p["primer_pago"]:
                p["primer_pago"] = r["fecha"]
            if not p["ultimo_pago"] or r["fecha"] > p["ultimo_pago"]:
                p["ultimo_pago"] = r["fecha"]

    persona_list = [
        {
            "identificacion": ident,
            "nombres": d["nombres"],
            "total_gastado": d["total_gastado"],
            "transacciones": d["transacciones"],
            "flyers": sorted(d["flyers"]),
            "primer_pago": d["primer_pago"],
            "ultimo_pago": d["ultimo_pago"],
        }
        for ident, d in sorted(personas.items(), key=lambda x: -x[1]["total_gastado"])
    ]
    return {"personas": persona_list, "total": len(persona_list)}


@app.get("/api/crm/data")
async def get_crm_data():
    data = _load_crm_data()
    if data is None:
        raise HTTPException(status_code=404, detail="CRM data file not found")
    return data


def formatear_fecha_simple(valor):
    if not valor:
        return "-"
    try:
        d = datetime.fromisoformat(valor.replace("Z", "+00:00")).astimezone(COL_TZ)
        meses = ["enero", "febrero", "marzo", "abril", "mayo", "junio",
                  "julio", "agosto", "septiembre", "octubre", "noviembre", "diciembre"]
        return f"{d.day} de {meses[d.month - 1]} de {d.year} a las {d.hour}:{d.minute:02d}"
    except Exception:
        return str(valor)


@app.get("/api/download")
async def download_xlsx():
    result = supabase.table("usuarios_funlidi").select("*").order("updated_at", desc=True).execute()
    rows = result.data

    wb = Workbook()
    ws = wb.active
    ws.title = "Usuarios FUNLIDI"

    headers = [
        "Usuario de Telegram", "Nombres Completos",
        "Correo Electronico", "Numero de Documento",
        "Fecha de Creacion", "Ultima Actualizacion",
    ]
    ws.append(headers)

    for r in rows:
        usuario = r.get("telegram_username")
        if usuario:
            usuario = "@" + usuario
        else:
            usuario = "-"
        ws.append([
            usuario,
            r.get("nombres_completos") or "-",
            r.get("correo_electronico") or "-",
            r.get("numero_documento") or "-",
            formatear_fecha_simple(r.get("created_at")),
            formatear_fecha_simple(r.get("updated_at")),
        ])

    from openpyxl.styles import Font, PatternFill
    header_fill = PatternFill(start_color="FFC107", end_color="FFC107", fill_type="solid")
    header_font = Font(bold=True, size=11)
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font

    for column in ws.columns:
        max_len = 0
        col_letter = column[0].column_letter
        for cell in column:
            try:
                val = str(cell.value) if cell.value else ""
                max_len = max(max_len, len(val))
            except Exception:
                pass
        ws.column_dimensions[col_letter].width = min(max_len + 4, 40)

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    hoy = datetime.now(COL_TZ)
    filename = f"Registros_FUNLIDI_{hoy.day:02d}-{hoy.month:02d}-{hoy.year}.xlsx"
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


# ========== AYUDAS HUMANITARIAS ENDPOINTS ==========

AYUDAS_TABLE = "ayudas_humanitarias"
AYUDAS_BENEF_TABLE = "ayudas_beneficiarios"
AYUDAS_PERSONAL_FIELDS = ["nombre", "dni", "pais", "ciudad", "pasaporte", "ocupacion", "telefono", "correo"]
AYUDAS_BANK_FIELDS = ["banco", "swift", "nbancaria", "tipocuenta"]
AYUDAS_ALL_FIELDS = AYUDAS_PERSONAL_FIELDS + AYUDAS_BANK_FIELDS


def _ayudas_estado(r):
    personal = all(r.get(f) and str(r.get(f, "")).strip() for f in AYUDAS_PERSONAL_FIELDS)
    if not personal:
        return "incompleto"
    bank = all(r.get(f) and str(r.get(f, "")).strip() for f in AYUDAS_BANK_FIELDS)
    if not bank:
        return "sin_banco"
    bank_real = all(
        str(r.get(f, "")).strip() not in ("", "N/A") for f in AYUDAS_BANK_FIELDS
    )
    if bank_real:
        return "completo"
    return "sin_banco"


@app.get("/api/ayudas/data")
async def get_ayudas_data():
    result = supabase_ayudas.table(AYUDAS_TABLE).select("*").order("updated_at", desc=True).execute()
    rows = result.data or []
    for r in rows:
        r["estado"] = _ayudas_estado(r)
        uid = r.get("telegram_user_id")
        if uid is not None:
            benef = supabase_ayudas.table(AYUDAS_BENEF_TABLE).select("*").eq("telegram_user_id", uid).order("beneficiary_number").execute()
            r["beneficiarios"] = benef.data or []
        else:
            r["beneficiarios"] = []
    return {"data": rows, "total": len(rows)}


@app.get("/api/ayudas/stats")
async def get_ayudas_stats():
    result = supabase_ayudas.table(AYUDAS_TABLE).select("*").execute()
    rows = result.data or []
    total = len(rows)
    completos = sum(1 for r in rows if _ayudas_estado(r) == "completo")
    sin_banco = sum(1 for r in rows if _ayudas_estado(r) == "sin_banco")
    incompletos = sum(1 for r in rows if _ayudas_estado(r) == "incompleto")
    paises = set()
    for r in rows:
        p = r.get("pais")
        if p and str(p).strip() and str(p).strip() != "VACIO":
            paises.add(str(p).strip().upper())
    ultima = max(
        (r.get("updated_at") or r.get("created_at") or "") for r in rows
    ) if rows else None

    ahora_col = datetime.now(COL_TZ)
    hoy_inicio_col = ahora_col.replace(hour=0, minute=0, second=0, microsecond=0)
    semana_inicio_col = hoy_inicio_col - timedelta(days=7)
    registros_hoy = 0
    registros_semana = 0
    for r in rows:
        c = r.get("created_at")
        if c:
            try:
                d = datetime.fromisoformat(c.replace("Z", "+00:00")).astimezone(COL_TZ)
                if d >= hoy_inicio_col:
                    registros_hoy += 1
                if d >= semana_inicio_col:
                    registros_semana += 1
            except Exception:
                pass

    benef_result = supabase_ayudas.table(AYUDAS_BENEF_TABLE).select("*").execute()
    total_benef = len(benef_result.data or [])

    paises_list = sorted(paises) if paises else []
    return {
        "total": total,
        "completos": completos,
        "sin_banco": sin_banco,
        "incompletos": incompletos,
        "ultima_actualizacion": ultima,
        "registros_hoy": registros_hoy,
        "registros_semana": registros_semana,
        "paises": paises_list,
        "total_beneficiarios": total_benef,
    }


@app.get("/api/ayudas/download")
async def download_ayudas_xlsx():
    result = supabase_ayudas.table(AYUDAS_TABLE).select("*").order("updated_at", desc=True).execute()
    rows = result.data or []

    wb = Workbook()
    ws = wb.active
    ws.title = "Ayudas Humanitarias"

    headers = [
        "Usuario Telegram", "Tratamiento Datos",
        "Nombres y Apellidos", "Cedula/DNI", "Pais", "Ciudad",
        "Pasaporte", "Ocupacion", "Telefono", "Correo",
        "Banco", "Swift", "Numero Cuenta", "Tipo Cuenta",
        "Fecha Creacion", "Ultima Actualizacion",
    ]
    ws.append(headers)

    for r in rows:
        usuario = r.get("telegram_username")
        if usuario:
            usuario = "@" + usuario
        else:
            usuario = "-"
        ws.append([
            usuario,
            "Si" if r.get("data_treatment_accepted") else "No",
            r.get("nombre") or "-", r.get("dni") or "-",
            r.get("pais") or "-", r.get("ciudad") or "-",
            r.get("pasaporte") or "-", r.get("ocupacion") or "-",
            r.get("telefono") or "-", r.get("correo") or "-",
            r.get("banco") or "-", r.get("swift") or "-",
            r.get("nbancaria") or "-", r.get("tipocuenta") or "-",
            formatear_fecha_simple(r.get("created_at")),
            formatear_fecha_simple(r.get("updated_at")),
        ])

    from openpyxl.styles import Font, PatternFill
    header_fill = PatternFill(start_color="64B5F6", end_color="64B5F6", fill_type="solid")
    header_font = Font(bold=True, size=11)
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font

    for column in ws.columns:
        max_len = 0
        col_letter = column[0].column_letter
        for cell in column:
            try:
                val = str(cell.value) if cell.value else ""
                max_len = max(max_len, len(val))
            except Exception:
                pass
        ws.column_dimensions[col_letter].width = min(max_len + 4, 40)

    ws2 = wb.create_sheet(title="Beneficiarios")
    benef_headers = [
        "Usuario Telegram", "Beneficiario #",
        "Nombres y Apellidos", "Cedula/DNI", "Pais", "Ciudad",
        "Pasaporte", "Ocupacion", "Telefono", "Correo",
    ]
    ws2.append(benef_headers)
    for r in rows:
        usuario = r.get("telegram_username")
        if usuario:
            usuario = "@" + usuario
        else:
            usuario = "-"
        uid = r.get("telegram_user_id")
        if uid is not None:
            benef_res = supabase_ayudas.table(AYUDAS_BENEF_TABLE).select("*").eq("telegram_user_id", uid).order("beneficiary_number").execute()
            for b in (benef_res.data or []):
                ws2.append([
                    usuario,
                    b.get("beneficiary_number", ""),
                    b.get("nombre") or "-", b.get("dni") or "-",
                    b.get("pais") or "-", b.get("ciudad") or "-",
                    b.get("pasaporte") or "-", b.get("ocupacion") or "-",
                    b.get("telefono") or "-", b.get("correo") or "-",
                ])
    from openpyxl.styles import Font, PatternFill
    header_fill2 = PatternFill(start_color="64B5F6", end_color="64B5F6", fill_type="solid")
    header_font2 = Font(bold=True, size=11)
    for cell in ws2[1]:
        cell.fill = header_fill2
        cell.font = header_font2
    for column in ws2.columns:
        max_len = 0
        col_letter = column[0].column_letter
        for cell in column:
            try:
                val = str(cell.value) if cell.value else ""
                max_len = max(max_len, len(val))
            except Exception:
                pass
        ws2.column_dimensions[col_letter].width = min(max_len + 4, 40)

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    hoy = datetime.now(COL_TZ)
    filename = f"Ayudas_Humanitarias_{hoy.day:02d}-{hoy.month:02d}-{hoy.year}.xlsx"
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


# ========== INVENTARIO ENDPOINTS ==========

INVENTARIO_TABLE = "inventario_adquisiciones"


@app.get("/api/inventario/data")
async def get_inventario_data():
    result = supabase_inventario.table(INVENTARIO_TABLE).select("*").order("updated_at", desc=True).execute()
    rows = result.data or []
    return {"data": rows, "total": len(rows)}


@app.get("/api/inventario/stats")
async def get_inventario_stats():
    result = supabase_inventario.table(INVENTARIO_TABLE).select("*").execute()
    rows = result.data or []
    total = len(rows)

    cajamicro_total = sum(to_int(r.get("cajamicro")) for r in rows)
    cajadinar_total = sum(to_int(r.get("cajadinar")) for r in rows)
    per_aleman_total = sum(to_int(r.get("per_aleman")) for r in rows)
    per_top_total = sum(to_int(r.get("per_top")) for r in rows)
    per_dragon_total = sum(to_int(r.get("per_dragon")) for r in rows)

    ultima = max(
        (r.get("updated_at") or r.get("created_at") or "") for r in rows
    ) if rows else None

    ahora_col = datetime.now(COL_TZ)
    hoy_inicio_col = ahora_col.replace(hour=0, minute=0, second=0, microsecond=0)
    semana_inicio_col = hoy_inicio_col - timedelta(days=7)
    registros_hoy = 0
    registros_semana = 0
    for r in rows:
        c = r.get("created_at") or r.get("updated_at")
        if c:
            try:
                d = datetime.fromisoformat(c.replace("Z", "+00:00")).astimezone(COL_TZ)
                if d >= hoy_inicio_col:
                    registros_hoy += 1
                if d >= semana_inicio_col:
                    registros_semana += 1
            except Exception:
                pass

    return {
        "total": total,
        "cajamicro_total": cajamicro_total,
        "cajadinar_total": cajadinar_total,
        "per_aleman_total": per_aleman_total,
        "per_top_total": per_top_total,
        "per_dragon_total": per_dragon_total,
        "ultima_actualizacion": ultima,
        "registros_hoy": registros_hoy,
        "registros_semana": registros_semana,
    }


def to_int(v):
    if not v:
        return 0
    try:
        return int(v)
    except Exception:
        return 0


@app.get("/api/inventario/download")
async def download_inventario_xlsx():
    result = supabase_inventario.table(INVENTARIO_TABLE).select("*").order("updated_at", desc=True).execute()
    rows = result.data or []

    wb = Workbook()
    ws = wb.active
    ws.title = "Inventario de Adquisiciones"

    headers = [
        "Usuario Telegram", "Nombres y Apellidos", "Cedula/DNI", "Pais",
        "Cajas Microlingotes", "Cajas Dinares", "Pergaminos Alemanes",
        "Pergaminos Nonillon", "Cajas Pergaminos Dragones",
        "Fecha Creacion", "Ultima Actualizacion",
    ]
    ws.append(headers)

    for r in rows:
        usuario = r.get("telegram_username")
        if usuario:
            usuario = "@" + usuario
        else:
            usuario = "-"
        ws.append([
            usuario,
            r.get("nombre") or "-",
            r.get("dni") or "-",
            r.get("pais") or "-",
            to_int(r.get("cajamicro")),
            to_int(r.get("cajadinar")),
            to_int(r.get("per_aleman")),
            to_int(r.get("per_top")),
            to_int(r.get("per_dragon")),
            formatear_fecha_simple(r.get("created_at") or r.get("updated_at")),
            formatear_fecha_simple(r.get("updated_at")),
        ])

    from openpyxl.styles import Font, PatternFill
    header_fill = PatternFill(start_color="81C784", end_color="81C784", fill_type="solid")
    header_font = Font(bold=True, size=11)
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font

    for column in ws.columns:
        max_len = 0
        col_letter = column[0].column_letter
        for cell in column:
            try:
                val = str(cell.value) if cell.value else ""
                max_len = max(max_len, len(val))
            except Exception:
                pass
        ws.column_dimensions[col_letter].width = min(max_len + 4, 40)

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    hoy = datetime.now(COL_TZ)
    filename = f"Inventario_Adquisiciones_{hoy.day:02d}-{hoy.month:02d}-{hoy.year}.xlsx"
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


# ========== INVENTARIO CRUD ENDPOINTS ==========


INVENTARIO_FIELDS = ["nombre", "dni", "pais", "cajamicro", "cajadinar", "per_aleman", "per_top", "per_dragon"]


def _next_negative_id():
    result = (
        supabase_inventario.table(INVENTARIO_TABLE)
        .select("telegram_user_id")
        .lt("telegram_user_id", 0)
        .order("telegram_user_id")
        .execute()
    )
    existing = [r["telegram_user_id"] for r in (result.data or [])]
    if not existing:
        return -1
    return min(existing) - 1


@app.post("/api/inventario/add")
async def add_inventario_entry(data: dict = Body(...)):
    telegram_username = (data.get("telegram_username") or "").strip()
    if not telegram_username:
        telegram_username = "MANUAL"

    nombre = (data.get("nombre") or "").strip()
    if not nombre:
        raise HTTPException(status_code=400, detail="El nombre es obligatorio")

    telegram_user_id = _next_negative_id()
    now = datetime.now(timezone.utc).isoformat()

    row = {
        "telegram_user_id": telegram_user_id,
        "telegram_username": telegram_username,
        "created_at": now,
        "updated_at": now,
    }
    for k in INVENTARIO_FIELDS:
        v = data.get(k)
        row[k] = str(v).strip() if v and str(v).strip() else None

    supabase_inventario.table(INVENTARIO_TABLE).upsert(row, on_conflict="telegram_user_id").execute()
    return {"success": True, "data": row}


@app.put("/api/inventario/edit/{telegram_user_id}")
async def edit_inventario_entry(telegram_user_id: int, data: dict = Body(...)):
    row = {}
    for k in INVENTARIO_FIELDS:
        if k in data:
            v = data[k]
            row[k] = str(v).strip() if v and str(v).strip() else None

    if "telegram_username" in data:
        v = data["telegram_username"]
        row["telegram_username"] = str(v).strip() if v and str(v).strip() else "MANUAL"

    if not row:
        raise HTTPException(status_code=400, detail="No fields to update")

    row["updated_at"] = datetime.now(timezone.utc).isoformat()

    supabase_inventario.table(INVENTARIO_TABLE).update(row).eq("telegram_user_id", telegram_user_id).execute()
    return {"success": True}


@app.delete("/api/inventario/delete/{telegram_user_id}")
async def delete_inventario_entry(telegram_user_id: int):
    supabase_inventario.table(INVENTARIO_TABLE).delete().eq("telegram_user_id", telegram_user_id).execute()
    return {"success": True}


# ========== CONSULTA ENDPOINT ==========


def _match_q(val, q):
    return val and q.lower() in str(val).lower()


def _norm(val):
    if not val:
        return ""
    return str(val).strip().lower()


@app.get("/api/consulta")
async def get_consulta(q: str = ""):
    q = q.strip()
    if len(q) < 2:
        return {"query": q, "total": 0, "persons": []}

    persons = {}

    def get_or_create(key):
        if key not in persons:
            persons[key] = {
                "name": "",
                "identifiers": {},
                "bot": {"exists": False},
                "ayudas": {"exists": False},
                "inventario": {"exists": False},
                "pagos": {"exists": False},
            }
        return persons[key]

    def link_record(rec, key_fields, source_name, extract_fn):
        def link_key(r):
            parts = []
            for k in key_fields:
                v = _norm(r.get(k))
                if v:
                    parts.append(f"{k}:{v}")
            return "|".join(parts) if parts else None

        keys = set()
        for field in key_fields:
            v = rec.get(field)
            if v is not None and str(v).strip():
                keys.add(("field:" + str(v).strip().lower(), source_name))

        if not keys:
            keys.add((source_name + ":" + str(id(rec)), source_name))

        for k, src in keys:
            p = get_or_create(k)
            extract_fn(p, rec)
            for other_k, other_src in keys:
                if other_k != k:
                    persons[other_k] = p

    # --- 1. Buscar en BOT (usuarios_funlidi) ---
    try:
        bot_result = supabase.table("usuarios_funlidi").select("*").execute()
        for r in (bot_result.data or []):
            if not any(_match_q(r.get(f), q) for f in ["nombres_completos", "numero_documento", "correo_electronico", "telegram_username"]):
                continue
            def extract_bot(p, rec):
                p["name"] = p["name"] or rec.get("nombres_completos") or ""
                p["identifiers"]["telegram_user_id"] = rec.get("telegram_user_id")
                p["identifiers"]["telegram_username"] = rec.get("telegram_username") or p["identifiers"].get("telegram_username")
                p["identifiers"]["dni"] = rec.get("numero_documento") or p["identifiers"].get("dni")
                p["identifiers"]["email"] = rec.get("correo_electronico") or p["identifiers"].get("email")
                p["bot"] = {
                    "exists": True,
                    "nombres_completos": rec.get("nombres_completos"),
                    "numero_documento": rec.get("numero_documento"),
                    "correo_electronico": rec.get("correo_electronico"),
                    "telegram_username": rec.get("telegram_username"),
                    "updated_at": rec.get("updated_at") or rec.get("created_at"),
                }
            link_record(r, ["telegram_user_id", "numero_documento", "correo_electronico"], "bot", extract_bot)
    except Exception:
        pass

    # --- 2. Buscar en AYUDAS (ayudas_humanitarias) ---
    try:
        ayudas_result = supabase_ayudas.table(AYUDAS_TABLE).select("*").execute()
        for r in (ayudas_result.data or []):
            if not any(_match_q(r.get(f), q) for f in ["nombre", "dni", "correo", "telegram_username"]):
                continue
            def extract_ayudas(p, rec):
                p["name"] = p["name"] or rec.get("nombre") or ""
                p["identifiers"]["telegram_user_id"] = rec.get("telegram_user_id") or p["identifiers"].get("telegram_user_id")
                p["identifiers"]["telegram_username"] = rec.get("telegram_username") or p["identifiers"].get("telegram_username")
                p["identifiers"]["dni"] = rec.get("dni") or p["identifiers"].get("dni")
                p["identifiers"]["email"] = rec.get("correo") or p["identifiers"].get("email")
                benef_count = 0
                uid = rec.get("telegram_user_id")
                if uid is not None:
                    try:
                        benef = supabase_ayudas.table(AYUDAS_BENEF_TABLE).select("*").eq("telegram_user_id", uid).execute()
                        benef_count = len(benef.data or [])
                    except Exception:
                        pass
                p["ayudas"] = {
                    "exists": True,
                    "nombre": rec.get("nombre"),
                    "dni": rec.get("dni"),
                    "pais": rec.get("pais"),
                    "estado": _ayudas_estado(rec),
                    "beneficiarios": benef_count,
                    "telefono": rec.get("telefono"),
                    "correo": rec.get("correo"),
                    "updated_at": rec.get("updated_at") or rec.get("created_at"),
                }
            link_record(r, ["telegram_user_id", "dni", "correo"], "ayudas", extract_ayudas)
    except Exception:
        pass

    # --- 3. Buscar en INVENTARIO (inventario_adquisiciones) ---
    try:
        inv_result = supabase_inventario.table(INVENTARIO_TABLE).select("*").execute()
        for r in (inv_result.data or []):
            if not any(_match_q(r.get(f), q) for f in ["nombre", "dni", "telegram_username"]):
                continue
            def extract_inventario(p, rec):
                p["name"] = p["name"] or rec.get("nombre") or ""
                p["identifiers"]["telegram_user_id"] = rec.get("telegram_user_id") or p["identifiers"].get("telegram_user_id")
                p["identifiers"]["telegram_username"] = rec.get("telegram_username") or p["identifiers"].get("telegram_username")
                p["identifiers"]["dni"] = rec.get("dni") or p["identifiers"].get("dni")
                p["inventario"] = {
                    "exists": True,
                    "nombre": rec.get("nombre"),
                    "dni": rec.get("dni"),
                    "pais": rec.get("pais"),
                    "materiales": {
                        "cajamicro": to_int(rec.get("cajamicro")),
                        "cajadinar": to_int(rec.get("cajadinar")),
                        "per_aleman": to_int(rec.get("per_aleman")),
                        "per_top": to_int(rec.get("per_top")),
                        "per_dragon": to_int(rec.get("per_dragon")),
                    },
                    "updated_at": rec.get("updated_at") or rec.get("created_at"),
                }
            link_record(r, ["telegram_user_id", "dni"], "inventario", extract_inventario)
    except Exception:
        pass

    # --- 4. Buscar en PAGOS (Excel) ---
    try:
        pagos_records = _load_pagos_data()
        for r in pagos_records:
            if not any(_match_q(r.get(f), q) for f in ["nombres", "identificacion"]):
                continue
            def extract_pagos(p, rec):
                p["name"] = p["name"] or rec.get("nombres") or ""
                p["identifiers"]["dni"] = rec.get("identificacion") or p["identifiers"].get("dni")
                pagos_data = p.get("pagos", {})
                if not pagos_data.get("exists"):
                    pagos_data["exists"] = True
                    pagos_data["transacciones"] = 0
                    pagos_data["total_cop"] = 0
                    pagos_data["detalles"] = []
                    pagos_data["ultimo_pago"] = ""
                pagos_data["transacciones"] += 1
                pagos_data["total_cop"] += int(rec.get("valor", 0))
                if rec.get("fecha") and rec["fecha"] > pagos_data.get("ultimo_pago", ""):
                    pagos_data["ultimo_pago"] = rec["fecha"]
                pagos_data["detalles"].append({
                    "fecha": rec.get("fecha", ""),
                    "flayer": rec.get("flayer", ""),
                    "valor": int(rec.get("valor", 0)),
                })
                p["pagos"] = pagos_data
            link_record(r, ["identificacion"], "pagos", extract_pagos)
    except Exception:
        pass

    # --- 5. Buscar en CRM (B. DATOS - FARLEY) ---
    try:
        crm_data = _load_crm_data()
        if crm_data and crm_data.get("members"):
            for m in crm_data["members"]:
                if not any(_match_q(m.get(f), q) for f in ["name", "cedula", "email", "telegram", "phone"]):
                    continue
                def extract_crm(p, rec):
                    p["name"] = p["name"] or rec.get("name") or ""
                    p["identifiers"]["dni"] = rec.get("cedula") or p["identifiers"].get("dni")
                    p["identifiers"]["email"] = rec.get("email") or p["identifiers"].get("email")
                    tags = set()
                    for pu in rec.get("purchases", []):
                        for t in pu.get("tags", []):
                            tags.add(t)
                    p["farley"] = {
                        "exists": True,
                        "name": rec.get("name"),
                        "cedula": rec.get("cedula"),
                        "email": rec.get("email"),
                        "telegram": rec.get("telegram"),
                        "phone": rec.get("phone"),
                        "city": rec.get("city"),
                        "department": rec.get("department"),
                        "country": rec.get("country"),
                        "purchase_count": rec.get("purchase_count", 0),
                        "total_spent": rec.get("total_spent", 0),
                        "categories": sorted(tags),
                        "purchases": sorted(rec.get("purchases", []), key=lambda x: x.get("date", ""), reverse=True)[:10],
                    }
                link_record(m, ["cedula", "email"], "farley", extract_crm)
    except Exception:
        pass

    # Deduplicate persons dict into a list
    seen = set()
    result_list = []
    for key, p in persons.items():
        pid = str(p["identifiers"])
        if pid in seen:
            continue
        seen.add(pid)
        # Sort pagos detalles newest first
        if p["pagos"]["exists"]:
            p["pagos"]["detalles"].sort(key=lambda x: x.get("fecha", ""), reverse=True)
            p["pagos"]["detalles"] = p["pagos"]["detalles"][:10]
        result_list.append(p)

    result_list.sort(key=lambda p: p["name"] or "")

    return {"query": q, "total": len(result_list), "persons": result_list}
